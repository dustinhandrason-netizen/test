from flask import Flask, render_template, request, redirect, url_for, flash
import os, json, base64, random, time, csv
from werkzeug.utils import secure_filename
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
import threading
from openpyxl import load_workbook   # for Excel reading
from xhtml2pdf import pisa
from pdf2docx import Converter

def pdf_to_docx(pdf_file, docx_file):
    cv = Converter(pdf_file)
    cv.convert(docx_file, start=0, end=None)
    cv.close()
    return docx_file

def generate_pdf_from_html(html_content, output_filename):
    with open(output_filename, "w+b") as result:
        pisa.CreatePDF(html_content, dest=result)
    return output_filename



app = Flask(__name__)
app.secret_key = "super_secret_key"

# Upload folder for credentials, token, recipient files
UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# Gmail API settings
SCOPES = ["https://www.googleapis.com/auth/gmail.send"]
CLIENT_SECRETS_FILE = os.path.join(UPLOAD_FOLDER, "credentials.json")
TOKEN_FILE = os.path.join(UPLOAD_FOLDER, "token.json")

# Allow HTTP (local dev, disable in production if using HTTPS)
os.environ["OAUTHLIB_INSECURE_TRANSPORT"] = "1"


# ----------- Helpers ----------
def save_credentials(creds):
    with open(TOKEN_FILE, "w") as token_file:
        token_file.write(creds.to_json())

def load_credentials():
    """Try to load valid credentials from token.json"""
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, "r") as token_file:
            try:
                data = json.load(token_file)
                creds = Credentials.from_authorized_user_info(data, SCOPES)
                if creds and creds.valid:
                    return creds
            except Exception as e:
                print("Invalid token.json:", e)
    return None  # force re-authorization


@app.route("/upload_credentials", methods=["GET", "POST"])
def upload_credentials():
    if request.method == "POST":
        if "credentials" not in request.files:
            flash("No credentials.json file uploaded")
            return redirect(request.url)

        credentials_file = request.files["credentials"]

        if credentials_file and credentials_file.filename.endswith(".json"):
            credentials_path = os.path.join(app.config["UPLOAD_FOLDER"], "credentials.json")
            credentials_file.save(credentials_path)
            flash("credentials.json uploaded successfully")

        return redirect(url_for("upload_credentials"))

    return render_template("upload_credentials.html")


def send_via_gmail(service, to, subject, body, is_html=False, attachment_path=None):
    message = MIMEMultipart("mixed")
    message["to"] = to
    message["subject"] = subject

    # Body
    alt_part = MIMEMultipart("alternative")
    if is_html:
        alt_part.attach(MIMEText(body, "html"))
    else:
        alt_part.attach(MIMEText(body, "plain"))
    message.attach(alt_part)

    # Attachment
    if attachment_path and os.path.exists(attachment_path):
        from email.mime.base import MIMEBase
        from email import encoders

        with open(attachment_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(attachment_path)}")
            message.attach(part)

    raw = base64.urlsafe_b64encode(message.as_bytes()).decode()
    return service.users().messages().send(userId="me", body={"raw": raw}).execute()


def extract_recipients(uploaded_file, manual_text):
    recipients = []

    # Manual text box entries
    if manual_text.strip():
        recipients.extend([r.strip() for r in manual_text.splitlines() if r.strip()])

    # Uploaded file entries
    if uploaded_file and uploaded_file.filename:
        filename = secure_filename(uploaded_file.filename)
        path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        uploaded_file.save(path)

        if filename.endswith(".csv"):
            with open(path, newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                header = next(reader, None)
                for row in reader:
                    if not row:
                        continue
                    if header and "email" in [h.lower() for h in header]:
                        email_index = [h.lower() for h in header].index("email")
                        recipients.append(row[email_index].strip())
                    else:
                        recipients.append(row[0].strip())

        elif filename.endswith((".xls", ".xlsx")):
            wb = load_workbook(path, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return recipients
            header = rows[0]
            for row in rows[1:]:
                if not row:
                    continue
                if header and any(h and str(h).lower() == "email" for h in header):
                    email_index = [i for i, h in enumerate(header) if h and str(h).lower() == "email"][0]
                    recipients.append(str(row[email_index]).strip())
                else:
                    recipients.append(str(row[0]).strip())
            wb.close()

    return recipients


# ----------- Routes ----------
@app.route("/")
def index():
    creds = load_credentials()
    if not creds or not creds.valid:
        return redirect(url_for("authorize"))
    return render_template("bulk.html")


@app.route("/authorize")
def authorize():
    flow = Flow.from_client_secrets_file(
        CLIENT_SECRETS_FILE,
        scopes=SCOPES,
        redirect_uri=url_for("oauth2callback", _external=True),
    )
    auth_url, state = flow.authorization_url(
        access_type="offline", include_granted_scopes="true", prompt="consent"
    )
    return redirect(auth_url)


@app.route("/oauth2callback")
def oauth2callback():
    flow = Flow.from_client_secrets_file(
        CLIENT_SECRETS_FILE,
        scopes=SCOPES,
        redirect_uri=url_for("oauth2callback", _external=True),
    )
    flow.fetch_token(authorization_response=request.url)
    creds = flow.credentials
    save_credentials(creds)
    flash("Google account authorized successfully!", "success")
    return redirect(url_for("index"))


@app.route("/send_bulk", methods=["POST"])
def send_bulk():
    creds = load_credentials()
    if not creds or not creds.valid:
        return redirect(url_for("authorize"))

    service = build("gmail", "v1", credentials=creds)

    recipients = extract_recipients(request.files.get("file"), request.form["recipients"])
    subjects = request.form["subjects"].splitlines()
    tfn = request.form["tfn"].splitlines()
    bodies = request.form["bodies"].split("===")
    pause = int(request.form["pause"])
    is_html = "html" in request.form
    send_as_pdf = "pdf" in request.form 
    send_as_docx = "docx" in request.form
    bodies_with_pdf=[
    "Kindly find the attached invoice in PDF format.",
    "The invoice has been attached as a PDF for your reference.",
    "Please review the attached invoice document (PDF).",
    "Attached is the invoice in PDF format for your records.",
    "The requested invoice is attached below in PDF.",
    "For your convenience, the invoice has been attached as a PDF file.",
    "You will find the invoice attached in PDF format.",
    "Attached please find the invoice (PDF) for your review.",
    "The invoice document has been included as a PDF attachment.",
    "Please see the invoice attached here in PDF format."
]   # <--- NEW CHECKBOX

    # Save uploaded attachment (if any)
    attachment_file = request.files.get("attachment")
    attachment_path = None
    if attachment_file and attachment_file.filename:
        filename = secure_filename(attachment_file.filename)
        attachment_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        attachment_file.save(attachment_path)

    def background_task():
        for idx, recipient in enumerate(recipients, 1):
            subject = random.choice(subjects)
            body = random.choice(bodies)
            tfns = random.choice(tfn)

            # Personalize
            personalized_body = body.strip().replace("#NAME#", recipient.strip())
            personalized_body = personalized_body.replace("#EMAIL#", recipient.strip())
            personalized_body = personalized_body.replace("#TFN#", tfns)

            attach_path = attachment_path  # default if user uploaded something

            try:
                # Generate PDF
                if send_as_pdf:
                    number = random.randint(0, 9999)
                    pdf_path = os.path.join(app.config["UPLOAD_FOLDER"], f"order_{number}.pdf")
                    generate_pdf_from_html(personalized_body, pdf_path)
                    personalized_body = random.choice(bodies_with_pdf)
                    attach_path = pdf_path

                # Generate DOCX
                if send_as_docx:
                    number = random.randint(0, 9999)
                    pdf_path = os.path.join(app.config["UPLOAD_FOLDER"], f"order_{number}.pdf")
                    generate_pdf_from_html(personalized_body, pdf_path)
                    number = random.randint(0, 9999)
                    docx_path = os.path.join(app.config["UPLOAD_FOLDER"], f"order_{number}.docx")
                    pdf_to_docx(pdf_path, docx_path)
                    personalized_body = random.choice(bodies_with_pdf)
                    attach_path = docx_path

                    # delete intermediate pdf (optional cleanup)
                    if os.path.exists(pdf_path):
                        os.remove(pdf_path)

                # Send
                msg = send_via_gmail(
                    service,
                    recipient.strip(),
                    subject,
                    personalized_body,
                    is_html,
                    attach_path
                )
                print(f"{idx}/{len(recipients)} Sent to {recipient} (ID: {msg['id']})")

            except Exception as e:
                print(f"{idx}/{len(recipients)} Failed {recipient}: {str(e)}")

            finally:
                # Delete attachment after sending
                if attach_path and os.path.exists(attach_path) and attach_path != attachment_path:
                    try:
                        os.remove(attach_path)
                    except Exception as cleanup_err:
                        print(f"Failed to delete {attach_path}: {cleanup_err}")

            time.sleep(pause)


    threading.Thread(target=background_task).start()
    return f"Sending {len(recipients)} emails in background!"



if __name__ == "__main__":
    app.run(debug=True)
